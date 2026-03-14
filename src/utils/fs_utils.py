import os
import sys
import shutil
from datetime import datetime, timedelta
from src.utils.config import Config

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def resolve_conflict(dst_dir, filename):
    dst_file = os.path.join(dst_dir, filename)
    if not os.path.exists(dst_file):
        return dst_file
    base, ext = os.path.splitext(filename)
    counter = 1
    while counter < 1000:
        new_name = f"{base}_{counter}{ext}"
        new_dst = os.path.join(dst_dir, new_name)
        if not os.path.exists(new_dst):
            return new_dst
        counter += 1
    return dst_file

def get_date_based_dirs(base_root=None, mode='create', photographer_name="贺志"):
    """
    mode: 'create' (default) -> for generating folders
          'import' -> for importing files (原片)
    """
    if base_root is None:
        base_root = Config.get_root_dir()
    
    today = datetime.now()
    year_str = today.strftime("%Y")
    month_str = f"{today.month:02d}月"
    day_str = f"{today.month:02d}{today.day:02d}"
    
    if mode == 'import':
        # 导入模式: 统一使用 '原片' 后缀
        # 格式: .../{YYYY}相片/{MM}月/{MMDD}原片
        photo_dir = os.path.join(base_root, f"{year_str}相片", month_str, f"{day_str}原片")
        # 格式: .../{YYYY}VR/{MM}月/{MMDD}原片
        vr_dir = os.path.join(base_root, f"{year_str}VR", month_str, f"{day_str}原片")
    else:
        # 创建模式: 相片使用 photographer_name, VR 无后缀
        # 格式: .../{YYYY}相片/{MM}月/{MMDD}{photographer_name}
        photo_dir = os.path.join(base_root, f"{year_str}相片", month_str, f"{day_str}{photographer_name}")
        # 格式: .../{YYYY}VR/{MM}月/{MMDD}
        vr_dir = os.path.join(base_root, f"{year_str}VR", month_str, day_str)
    
    return [photo_dir, vr_dir]

def copy_yesterday_excel_to_today(base_root=None, photographer_name="贺志"):
    if base_root is None:
        base_root = Config.get_root_dir()
    try:
        today = datetime.now()
        today_dirs = get_date_based_dirs(base_root=base_root, mode='create', photographer_name=photographer_name)
        today_photo_dir = today_dirs[0]
        today_day_str = f"{today.month:02d}{today.day:02d}"
        os.makedirs(today_photo_dir, exist_ok=True)

        today_excel = os.path.join(
            today_photo_dir,
            f"{today_day_str}{photographer_name}.xlsx",
        )
        # 如果今天的 Excel 已经存在，不覆盖
        if os.path.exists(today_excel):
            return True

        source_excel = None
        for delta in range(1, 366):
            candidate = today - timedelta(days=delta)
            y_year_str = candidate.strftime("%Y")
            y_month_str = f"{candidate.month:02d}月"
            y_day_str = f"{candidate.month:02d}{candidate.day:02d}"
            candidate_dir = os.path.join(
                base_root,
                f"{y_year_str}相片",
                y_month_str,
                f"{y_day_str}{photographer_name}",
            )
            candidate_excel = os.path.join(
                candidate_dir,
                f"{y_day_str}{photographer_name}.xlsx",
            )
            if os.path.exists(candidate_excel):
                source_excel = candidate_excel
                break

        if not source_excel:
            return False

        shutil.copy2(source_excel, today_excel)
        return True
    except Exception:
        return False

def is_same_device(src, dst):
    try:
        if not os.path.exists(src) or not os.path.exists(dst):
            # If destination doesn't exist, check parent
            dst_check = dst if os.path.exists(dst) else os.path.dirname(dst)
            # If src doesn't exist, can't check, assume False
            if not os.path.exists(src):
                return False
            return os.stat(src).st_dev == os.stat(dst_check).st_dev
        return os.stat(src).st_dev == os.stat(dst).st_dev
    except Exception:
        return False

def _parse_shoot_line(line):
    raw = (line or "").strip()
    if not raw:
        return None
    raw = raw.replace("　", " ")
    raw = raw.strip()
    import re
    
    # Parse sequence number from the start of the line
    seq = None
    match = re.match(r'^\s*(\d+)\s*[\.、．。]?\s*', raw)
    if match:
        seq = int(match.group(1))
        raw = raw[match.end():] # Remove the sequence part
    else:
        # Fallback stripping if no group capture but still some digits (should be covered by above, but safe to keep regex clean)
        raw = re.sub(r'^\s*\d+\s*[\.、．。]?\s*', '', raw)

    parts = [p for p in raw.split() if p]
    if not parts:
        return None
        
    name = parts[0]
    hs = None
    hs_idx = None
    for idx, p in enumerate(parts):
        if p.startswith("HS") and len(p) >= 4:
            hs = p
            hs_idx = idx
            break
            
    if hs and hs_idx is not None and hs_idx + 1 < len(parts):
        direction = parts[-1] if len(parts) >= 2 else ""
        room = parts[-2] if len(parts) >= 3 else ""
        store = parts[hs_idx + 1] if hs_idx + 1 < len(parts) else ""
        addr_parts = parts[hs_idx + 2:-2] if len(parts) >= 3 else []
        address = " ".join(addr_parts).strip()
    else:
        hs = ""
        # If parts only contains name (len=1), everything else is empty
        direction = parts[-1] if len(parts) >= 2 else ""
        room = parts[-2] if len(parts) >= 3 else ""
        
        mid = parts[1:-2] if len(parts) >= 4 else parts[1:-1]

        if len(mid) >= 2:
            store = mid[0]
            address = " ".join(mid[1:]).strip()
        elif len(mid) == 1:
            store = "" 
            address = mid[0]
        else:
            store = ""
            address = ""
            
    return {
        "name": name,
        "hs": hs,
        "store": store,
        "address": address,
        "room": room,
        "direction": direction,
        "seq": seq
    }

def _find_header_row_and_map(ws):
    wanted = {
        "photographer": ["摄影师"],
        "name": ["接单人", "经办人", "经纪人"],
        "hs": ["原房源号", "房源号", "原房源编号", "编号"],
        "store": ["接单人所属门店", "所属门店", "门店"],
        "address": ["房源地址", "小区", "楼盘名称", "楼盘", "房勘社区"],
        "room": ["房号", "房源房号", "门牌号"],
        "direction": ["入户门", "朝向"],
        "shoot_date": ["拍摄时间", "拍摄日期", "日期"],
        "seq": ["序号"],
        "monthly_count": ["当月套数", "套数"],
    }

    def norm(v):
        if v is None:
            return ""
        return str(v).strip()

    best_row = None
    best_score = -1
    best_map = {}

    max_scan_rows = min(20, ws.max_row or 20)
    for r in range(1, max_scan_rows + 1):
        row_map = {}
        score = 0
        for c in range(1, (ws.max_column or 1) + 1):
            v = norm(ws.cell(row=r, column=c).value)
            if not v:
                continue
            for key, aliases in wanted.items():
                if key in row_map:
                    continue
                for a in aliases:
                    if a in v:
                        row_map[key] = c
                        score += 1
                        break
        if score > best_score:
            best_score = score
            best_row = r
            best_map = row_map

    return best_row, best_map, best_score

def _pick_target_sheet_and_header(wb):
    best = None
    for ws in wb.worksheets:
        header_row, header_map, header_score = _find_header_row_and_map(ws)
        if not header_row:
            header_row = 1
            header_map = {}
            header_score = -1
        hs_col = _find_hs_column(ws, header_row, header_map)
        score = header_score
        if hs_col:
            score += 2
        candidate = (score, ws, header_row, header_map, hs_col)
        if best is None or candidate[0] > best[0]:
            best = candidate
    return best

def _find_hs_column(ws, header_row, header_map):
    hs_col = header_map.get("hs")
    if hs_col:
        return hs_col
    for c in range(1, (ws.max_column or 1) + 1):
        v = ws.cell(row=header_row, column=c).value
        if v and "HS" in str(v):
            return c
    scan_to = min(ws.max_row or 1, header_row + 50)
    for c in range(1, (ws.max_column or 1) + 1):
        for r in range(header_row + 1, scan_to + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.startswith("HS"):
                return c
    return None

def update_today_excel_from_folder_names(folder_names, base_root=None, max_backtrack_days=365, photographer_name="贺志"):
    if base_root is None:
        base_root = Config.get_root_dir()

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font
    except Exception:
        return 0, 0, "缺少 Excel 写入依赖（openpyxl）"

    today = datetime.now()
    today_dirs = get_date_based_dirs(base_root=base_root, mode="create", photographer_name=photographer_name)
    today_photo_dir = today_dirs[0]
    today_day_str = f"{today.month:02d}{today.day:02d}"
    today_excel = os.path.join(today_photo_dir, f"{today_day_str}{photographer_name}.xlsx")

    if not os.path.exists(today_excel):
        for delta in range(1, max_backtrack_days + 1):
            candidate = today - timedelta(days=delta)
            y_year_str = candidate.strftime("%Y")
            y_month_str = f"{candidate.month:02d}月"
            y_day_str = f"{candidate.month:02d}{candidate.day:02d}"
            candidate_dir = os.path.join(
                base_root,
                f"{y_year_str}相片",
                y_month_str,
                f"{y_day_str}{photographer_name}",
            )
            candidate_excel = os.path.join(candidate_dir, f"{y_day_str}{photographer_name}.xlsx")
            if os.path.exists(candidate_excel):
                os.makedirs(today_photo_dir, exist_ok=True)
                shutil.copy2(candidate_excel, today_excel)
                break

    if not os.path.exists(today_excel):
        return 0, 0, "未找到今日 Excel 文件，且无法从历史回溯复制"

    try:
        wb = load_workbook(today_excel)
    except Exception as e:
        return 0, 0, f"打开 Excel 失败：{e}"

    picked = _pick_target_sheet_and_header(wb)
    if not picked:
        wb.close()
        return 0, 0, "未找到可写入的工作表"

    _, ws, header_row, header_map, hs_col = picked
    if not hs_col:
        wb.close()
        return 0, 0, "未识别到 HS 列（可能表头行超出扫描范围或表头被合并）"

    # Find the real last data row by scanning HS column
    last_data_row = header_row
    max_scan_r = ws.max_row or header_row
    for r in range(header_row + 1, max_scan_r + 1):
        v_hs = ws.cell(row=r, column=hs_col).value
        v_seq = None
        seq_col = header_map.get("seq")
        if seq_col:
            v_seq = ws.cell(row=r, column=seq_col).value
            
        if (v_hs and str(v_hs).strip()) or (v_seq and str(v_seq).strip()):
            last_data_row = r
            
    # Calculate next monthly count (still need to auto-increment this)
    mc_col = header_map.get("monthly_count")
    mc_next = None
    if mc_col:
        max_mc = 0
        for r in range(last_data_row, header_row, -1):
            v = ws.cell(row=r, column=mc_col).value
            try:
                n = int(str(v).strip())
                max_mc = n
                break
            except Exception:
                continue
        mc_next = max_mc + 1

    date_col = header_map.get("shoot_date")
    date_str = f"{today.month}月{today.day}日"

    # Styles
    # Old Data: Fill C6EFCE (Light Green), Font 006100 (Dark Green)
    old_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    old_font_color = "006100"

    # New Data HS: Fill FFEB9C (Yellow), Font 9C5700 (Dark Brown)
    new_hs_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    new_hs_font_color = "9C5700"
    
    # New Data Other: No Fill (White), Font 000000 (Black)
    new_default_fill = PatternFill(fill_type=None)
    new_default_font_color = "000000"

    # Apply styles ONLY to the LAST found date (that is not today)
    # This optimization avoids scanning/coloring thousands of old rows repeatedly.
    if date_col:
        # Find the last date string that is NOT today
        last_diff_date_str = None
        # Scan upwards from last data row
        for r in range(last_data_row, header_row, -1):
            cell_date = ws.cell(row=r, column=date_col).value
            s_date = str(cell_date).strip()
            if s_date and s_date != date_str:
                last_diff_date_str = s_date
                break
        
        # If we found a "last different date", only color rows with THAT date
        if last_diff_date_str:
            # We can optimize scanning by only looking at recent rows, but safety first: scan down
            # Actually, since we want to color *all* rows of that last date (e.g. yesterday),
            # we should scan. To avoid scanning 5000 rows, maybe we can assume they are near the end?
            # But "yesterday" rows might be 50 rows up.
            # Let's scan from last_data_row upwards until date changes again?
            # Yes, data is usually chronological.
            
            # Scan upwards from last_data_row to color "yesterday's" data
            for r in range(last_data_row, header_row, -1):
                cell_date = ws.cell(row=r, column=date_col).value
                s_date = str(cell_date).strip()
                
                # If we hit a date that is neither today nor the target "last diff date", stop.
                # (Assuming sorted by date. If not sorted, this might miss some, but Excel usually is)
                if s_date != date_str and s_date != last_diff_date_str:
                    # We moved past the "last batch", so stop to save time
                    break
                
                if s_date == last_diff_date_str:
                    # Apply Old Data style
                    max_c = ws.max_column or 1
                    for c in range(1, max_c + 1):
                        # SKIP columns K, L, M, N (indices 11, 12, 13, 14)
                        if c in [11, 12, 13, 14]:
                            continue
                            
                        cell = ws.cell(row=r, column=c)
                        # Only apply if not already colored? No, user wants to ensure it's green.
                        # But to save write time, check if already green? 
                        # Accessing .fill is fast, comparing objects might be tricky.
                        # Let's just write. It's only ~20-50 rows usually.
                        
                        current_font = cell.font
                        cell.fill = old_fill
                        cell.font = Font(name=current_font.name, size=current_font.size, 
                                       bold=current_font.bold, italic=current_font.italic, 
                                       color=old_font_color)
                
                # If s_date == date_str (today), we might want to ensure it's New Style?
                # The prompt says "re-run" might happen. 
                # Yes, if user runs twice today, the first run's rows are now "existing" but have today's date.
                elif s_date == date_str:
                     max_c = ws.max_column or 1
                     for c in range(1, max_c + 1):
                        cell = ws.cell(row=r, column=c)
                        current_font = cell.font
                        if c == hs_col:
                            cell.fill = new_hs_fill
                            cell.font = Font(name=current_font.name, size=current_font.size, 
                                           bold=current_font.bold, italic=current_font.italic, 
                                           color=new_hs_font_color)
                        else:
                            cell.fill = new_default_fill
                            cell.font = Font(name=current_font.name, size=current_font.size, 
                                           bold=current_font.bold, italic=current_font.italic, 
                                           color=new_default_font_color)

    added = 0
    skipped = 0
    skipped_reasons = []
    write_columns = {
        "photographer": header_map.get("photographer"),
        "name": header_map.get("name"),
        "hs": hs_col,
        "store": header_map.get("store"),
        "address": header_map.get("address"),
        "room": header_map.get("room"),
        "direction": header_map.get("direction"),
        "shoot_date": date_col,
        "seq": seq_col,
        "monthly_count": mc_col,
    }

    from copy import copy as _copy

    def copy_row_style(src_row, dst_row):
        if src_row in ws.row_dimensions:
            ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
            
        max_c = ws.max_column or 1
        for c in range(1, max_c + 1):
            src_cell = ws.cell(row=src_row, column=c)
            dst_cell = ws.cell(row=dst_row, column=c)
            if src_cell.has_style:
                dst_cell._style = _copy(src_cell._style)
                dst_cell.font = _copy(src_cell.font)
                dst_cell.border = _copy(src_cell.border)
                # Do NOT copy fill, we set it explicitly
                # dst_cell.fill = _copy(src_cell.fill) 
                dst_cell.number_format = src_cell.number_format
                dst_cell.protection = _copy(src_cell.protection)
                dst_cell.alignment = _copy(src_cell.alignment)

    base_style_row = last_data_row if last_data_row > header_row else header_row + 1
    if base_style_row > (ws.max_row or 0):
        base_style_row = header_row

    for line in folder_names:
        parsed = _parse_shoot_line(line)
        if not parsed:
            skipped += 1
            skipped_reasons.append(f"解析失败: {line}")
            continue

        new_row = last_data_row + 1
        copy_row_style(base_style_row, new_row)

        # Apply New Data Styles
        max_c = ws.max_column or 1
        for c in range(1, max_c + 1):
            cell = ws.cell(row=new_row, column=c)
            current_font = cell.font
            
            if c == hs_col:
                cell.fill = new_hs_fill
                cell.font = Font(name=current_font.name, size=current_font.size, 
                               bold=current_font.bold, italic=current_font.italic, 
                               color=new_hs_font_color)
            else:
                cell.fill = new_default_fill
                cell.font = Font(name=current_font.name, size=current_font.size, 
                               bold=current_font.bold, italic=current_font.italic, 
                               color=new_default_font_color)

        if write_columns["photographer"]:
            ws.cell(row=new_row, column=write_columns["photographer"]).value = photographer_name
        if write_columns["name"]:
            ws.cell(row=new_row, column=write_columns["name"]).value = parsed["name"]
        ws.cell(row=new_row, column=write_columns["hs"]).value = parsed["hs"]
        if write_columns["store"]:
            ws.cell(row=new_row, column=write_columns["store"]).value = parsed["store"]
        if write_columns["address"]:
            ws.cell(row=new_row, column=write_columns["address"]).value = parsed["address"]
        if write_columns["room"]:
            ws.cell(row=new_row, column=write_columns["room"]).value = parsed["room"]
        if write_columns["direction"]:
            ws.cell(row=new_row, column=write_columns["direction"]).value = parsed["direction"]
        if write_columns["shoot_date"]:
            ws.cell(row=new_row, column=write_columns["shoot_date"]).value = date_str
            
        if write_columns["seq"]:
            if parsed.get("seq") is not None:
                ws.cell(row=new_row, column=write_columns["seq"]).value = parsed["seq"]
            
        if write_columns["monthly_count"] and mc_next is not None:
            ws.cell(row=new_row, column=write_columns["monthly_count"]).value = mc_next
            mc_next += 1

        last_data_row = new_row
        added += 1

    try:
        wb.save(today_excel)
    except PermissionError:
        wb.close()
        return 0, 0, "保存 Excel 失败：文件可能正被打开占用，请关闭后再试"
    except Exception as e:
        wb.close()
        return 0, 0, f"保存 Excel 失败：{e}"

    wb.close()
    
    msg = None
    if added == 0 and skipped > 0:
        detail = "; ".join(skipped_reasons[:2])
        msg = f"未新增记录。跳过 {skipped} 条。详情: {detail}"
        
    return added, skipped, msg
